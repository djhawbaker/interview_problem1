#!/usr/bin/env python
import pickle
import time
import argparse

from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import RED

def parseArgs():
    """Define and parse the input args
    
    Returns: 
        Processed args
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("file", type=str, help="Pickle file to convert to Excel format")
    parser.add_argument("-o", "--output", type=str, help="Output Excel file. Default='output.xlsx'")
    parser.set_defaults(output="output.xlsx")

    return parser.parse_args()


class WorkbookCreator():
    """Parses input pickle data and outputs an Excel workbook"""
    
    def __init__(self):
        """Constructor - Creates the Excel workbook"""
        self.wb = Workbook()
        self.ws = self.wb.active
        

    def setTitles(self):
        """Set title names and formatting
        Sets the sheet and column titles
        Bolds the titles
        """
        self.ws.title = "Task Data"

        self.ws['A1'] = "Task Name" # content
        self.ws['B1'] = "Set Part Name" # entity[code]
        self.ws['C1'] = "Parent Build Name" # sg_parent_build
        self.ws['D1'] = "Start Date" # start_date
        self.ws['E1'] = "End Date" # due_date

        # Set title formatting
        for col in range(1, self.ws.max_column+1):
            self.ws.cell(row=1, column=col).font = Font(bold=True)


    def setColumnWidth(self):
        """Set the column width to fit the widest cell
        For readability:
            Set max width of 80 characters
            Pad the end with a small amount of whitespace
        """
        for col in self.ws.columns:
            length = max(len(str(cell.value)) for cell in col)
            length += 2

            if (length > 80):
                length = 80

            self.ws.column_dimensions[col[0].column].width = length


    def parseEntry(self, entry):
        """Parses the entry and adds it to the next row of the worksheet

        Args:
            entry -- The current data entry object to parse
        """
        # Get the next row of the worksheet to populate
        nextRow = self.ws.max_row + 1

        # Verify the values exist before setting them to their cell
        # Task Name
        self.ws.cell(row=nextRow, 
                     column=1, 
                     value=entry['content'] if entry['content'] else '')

        # Set Part Name
        self.ws.cell(row=nextRow, 
                     column=2, 
                     value=entry['entity']['code'][0] if entry['entity'] else '')

        # Parent Build Name
        self.ws.cell(row=nextRow, 
                     column=3, 
                     value=entry['sg_parent_build']['code'] if entry['sg_parent_build'] else '')

        # Start Date
        self.ws.cell(row=nextRow, 
                     column=4, 
                     value=entry['start_date'] if entry['start_date'] else '')

        # End Date
        self.ws.cell(row=nextRow, 
                     column=5, 
                     value=entry['due_date'] if entry['due_date'] else '')

        self.addRedFill(entry, nextRow)


    def addRedFill(self, entry, row):
        """Adds a red background fill to the row if the entry is past due

        Args:
            entry -- The current data entry object to parse
            row -- The current row in the worksheet
        """
        # Define red pattern fill to use
        red = PatternFill(start_color=RED, end_color=RED, fill_type='solid')
         
        # If due date not entered, assume not past due
        if (entry['due_date']):
            due_date = time.strptime(entry['due_date'], '%Y-%m-%d')

            if (due_date < time.localtime()):
                for col in range(1, self.ws.max_column+1):
                    self.ws.cell(row=row, column=col).fill = red


    def importAndParse(self, pickle_file):
        """Imports and parses the data from the file
        Sorts the input data by the start date before parsing it
        into the Excel file

        Args:
            pickle_file -- name of input pickle file
        """
        # Import data
        with open(pickle_file, 'rb') as f:
            data = pickle.load(f)
        
        # Sort and parse data
        for entry in sorted(data, key=lambda x: x['start_date']):
            self.parseEntry(entry)


    def saveFile(self, output):
        """Saves the workbook to the output file

        Args:
            output -- name of output file
        """
        self.wb.save(output)


def main():
    # Parse the command line arguments
    args = parseArgs()
    pickle_file = args.file

    # Instantiate the workbook class
    wb = WorkbookCreator()

    # Set sheet and column title formatting
    wb.setTitles()

    # Import the pickle file and parse the contents
    wb.importAndParse(pickle_file)

    # Update the column widths for readability
    wb.setColumnWidth()

    # Save the Excel file
    wb.saveFile(args.output)


if __name__ == '__main__':
    main()
